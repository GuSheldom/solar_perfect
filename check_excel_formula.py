#!/usr/bin/env python3
"""检查Excel文件中的公式"""

import openpyxl

print("="*70)
print("检查Excel文件中POA列的公式")
print("="*70)

excel_file = '【1117版本】eiwomple完美收益模型修改.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    ws = wb.active
    
    print(f"\n工作表名: {ws.title}")
    print(f"最大行数: {ws.max_row}")
    print(f"最大列数: {ws.max_column}")
    
    # 读取标题行
    print("\n标题行（第1行）:")
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        cell = ws.cell(1, col)
        headers.append(cell.value)
        if col <= 10:
            print(f"  列{col} ({chr(64+col)}): {cell.value}")
    
    # 找到POA和PV功率的列
    poa_col = None
    pv_col = None
    
    for i, header in enumerate(headers, 1):
        if header and 'POA' in str(header):
            poa_col = i
        if header and 'PV功率' in str(header):
            pv_col = i
    
    print(f"\nPOA列: 第{poa_col}列 ({chr(64+poa_col) if poa_col else 'N/A'})")
    print(f"PV功率列: 第{pv_col}列 ({chr(64+pv_col) if pv_col else 'N/A'})")
    
    # 检查第2行（第一行数据）的POA单元格
    if poa_col:
        print("\n" + "="*70)
        print(f"检查POA列的前5行数据")
        print("="*70)
        
        for row in range(2, 7):
            cell = ws.cell(row, poa_col)
            print(f"\n第{row}行 POA单元格 ({chr(64+poa_col)}{row}):")
            print(f"  值: {cell.value}")
            print(f"  类型: {type(cell.value).__name__}")
            
            # 检查是否有公式
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"  数据类型: 公式")
            
            # 尝试读取公式（如果存在）
            try:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  公式: {cell.value}")
            except:
                pass
    
    # 检查PV功率列
    if pv_col:
        print("\n" + "="*70)
        print(f"检查PV功率列的前5行数据")
        print("="*70)
        
        for row in range(2, 7):
            cell = ws.cell(row, pv_col)
            print(f"\n第{row}行 PV功率单元格 ({chr(64+pv_col)}{row}):")
            print(f"  值: {cell.value}")
            print(f"  类型: {type(cell.value).__name__}")
            
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                print(f"  数据类型: 公式")
            
            try:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  公式: {cell.value}")
            except:
                pass
    
    wb.close()
    
    # 尝试以公式模式重新加载
    print("\n" + "="*70)
    print("尝试读取公式（公式模式）")
    print("="*70)
    
    wb_formula = openpyxl.load_workbook(excel_file, data_only=False)
    ws_formula = wb_formula.active
    
    if poa_col:
        cell = ws_formula.cell(2, poa_col)
        print(f"\nPOA (B2) 公式检查:")
        print(f"  cell.value: {cell.value}")
        
        # 检查是否是公式
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  ✓ 这是一个公式: {cell.value}")
        else:
            print(f"  ✗ 这不是公式，是直接数值")
    
    wb_formula.close()
    
except Exception as e:
    print(f"\n错误: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*70)


